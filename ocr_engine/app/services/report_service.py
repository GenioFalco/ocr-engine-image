"""
Ежедневный отчёт: собирает данные за сутки, строит Excel (2 листа),
отправляет на почту через SMTP.

Лист 1 — Сводка:  кол-во сканирований, токены, стоимость, оценки.
Лист 2 — Детали:  каждый документ с пользователем, оценкой, комментарием.
"""
import io
import json
import logging
import smtplib
from datetime import datetime, timedelta, timezone
from email.mime.application import MIMEApplication
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session

from app.config.settings import settings
from app.db.base import SessionLocal
from app.models.models import ProcessingJob, JobFeedback, Log, User
from app.services.settings_service import get_setting

logger = logging.getLogger(__name__)

# ── Цвета ──────────────────────────────────────────────────────────────────────
CLR_HEADER   = "1E293B"   # тёмно-синий заголовок
CLR_GOOD     = "DCFCE7"   # зелёный фон — «хорошо»
CLR_BAD      = "FEE2E2"   # красный фон — «плохо»
CLR_SUBHEAD  = "F1F5F9"   # серый подзаголовок


def _thin_border():
    side = Side(style="thin", color="CBD5E1")
    return Border(left=side, right=side, top=side, bottom=side)


def _header_font():
    return Font(color="FFFFFF", bold=True, size=10)


def _parse_token_log(msg: str):
    if not msg:
        return 0, 0, 0
    try:
        d = json.loads(msg)
        if isinstance(d, dict):
            t   = d.get("total", 0) or 0
            inp = d.get("input",  0) or 0
            out = d.get("output", 0) or 0
            return t, inp, out
    except Exception:
        pass
    try:
        v = int(msg)
        return v, 0, 0
    except Exception:
        return 0, 0, 0


def build_daily_report(db: Session, report_date: datetime | None = None) -> bytes:
    """
    Строит Excel и возвращает bytes.
    report_date — дата отчёта (UTC). По умолчанию — вчера.
    """
    if report_date is None:
        report_date = datetime.now(timezone.utc) - timedelta(days=1)

    day_start = report_date.replace(hour=0, minute=0, second=0, microsecond=0, tzinfo=timezone.utc)
    day_end   = day_start + timedelta(days=1)

    # ── Данные из БД ──────────────────────────────────────────────────────────
    jobs = (
        db.query(ProcessingJob)
        .filter(ProcessingJob.created_at >= day_start, ProcessingJob.created_at < day_end)
        .all()
    )
    job_ids = [j.id for j in jobs]

    feedbacks = {}
    if job_ids:
        for fb in db.query(JobFeedback).filter(JobFeedback.job_id.in_(job_ids)).all():
            feedbacks[fb.job_id] = fb

    # Токены за сутки
    token_logs = (
        db.query(Log)
        .filter(Log.stage == "TOKENS_USED", Log.created_at >= day_start, Log.created_at < day_end)
        .all()
    )
    total_tokens = total_input = total_output = 0
    for row in token_logs:
        t, inp, out = _parse_token_log(row.message)
        total_tokens += t; total_input += inp; total_output += out

    price_input_1m  = float(get_setting(db, "price_input_per_1m")  or 0.21)
    price_output_1m = float(get_setting(db, "price_output_per_1m") or 0.63)
    if total_input or total_output:
        cost_usd = total_input * price_input_1m / 1_000_000 + total_output * price_output_1m / 1_000_000
    else:
        blended = 0.75 * price_input_1m + 0.25 * price_output_1m
        cost_usd = total_tokens * blended / 1_000_000

    good_cnt = sum(1 for fb in feedbacks.values() if fb.rating >= 4)
    bad_cnt  = sum(1 for fb in feedbacks.values() if fb.rating <= 2)
    rated    = len(feedbacks)
    avg_rating = round(sum(fb.rating for fb in feedbacks.values()) / rated, 1) if rated else None

    users_map = {u.id: u for u in db.query(User).all()}

    # ── Workbook ──────────────────────────────────────────────────────────────
    wb = openpyxl.Workbook()

    # ══════════════════════════════════════════════════════════════════════════
    # Лист 1 — Сводка
    # ══════════════════════════════════════════════════════════════════════════
    ws1 = wb.active
    ws1.title = "Сводка"
    ws1.column_dimensions["A"].width = 35
    ws1.column_dimensions["B"].width = 22

    def _header_row(row, text):
        cell = ws1.cell(row=row, column=1, value=text)
        cell.font = Font(bold=True, color="FFFFFF", size=11)
        cell.fill = PatternFill("solid", fgColor=CLR_HEADER)
        cell.alignment = Alignment(horizontal="center")
        ws1.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)

    def _kv(row, key, value):
        c1 = ws1.cell(row=row, column=1, value=key)
        c2 = ws1.cell(row=row, column=2, value=value)
        c1.font = Font(bold=True, size=10)
        c2.alignment = Alignment(horizontal="right")
        c2.font = Font(size=10)
        c1.border = c2.border = _thin_border()

    r = 1
    _header_row(r, f"Ежедневный отчёт OCR Engine — {day_start.strftime('%d.%m.%Y')}")
    r += 1
    _header_row(r, "Статистика за сутки")
    r += 1
    _kv(r, "Дата отчёта", day_start.strftime("%d.%m.%Y")); r += 1
    _kv(r, "Всего сканирований",       len(jobs));          r += 1
    _kv(r, "Успешно (done)",           sum(1 for j in jobs if j.status == "done")); r += 1
    _kv(r, "Ошибки (failed)",          sum(1 for j in jobs if j.status == "failed")); r += 1
    _kv(r, "В работе / ожидании",      sum(1 for j in jobs if j.status in ("processing","pending"))); r += 1

    r += 1
    _header_row(r, "Токены и стоимость")
    r += 1
    _kv(r, "Всего токенов",             total_tokens);               r += 1
    _kv(r, "Input токены",              total_input);                r += 1
    _kv(r, "Output токены",             total_output);               r += 1
    _kv(r, "Стоимость (USD)",           f"${cost_usd:.6f}");         r += 1
    _kv(r, "Цена input ($/1M)",         f"${price_input_1m}");       r += 1
    _kv(r, "Цена output ($/1M)",        f"${price_output_1m}");      r += 1

    r += 1
    _header_row(r, "Качество распознавания (оценки пользователей)")
    r += 1
    _kv(r, "Оценено документов",        rated);                      r += 1
    c_good = ws1.cell(row=r, column=1, value="👍 Хорошо"); r_good = r
    _kv(r, "👍 Хорошо",                good_cnt);
    ws1.cell(row=r, column=1).fill = PatternFill("solid", fgColor=CLR_GOOD); r += 1
    _kv(r, "👎 Плохо",                 bad_cnt)
    ws1.cell(row=r, column=1).fill = PatternFill("solid", fgColor=CLR_BAD);  r += 1
    _kv(r, "Средняя оценка (1-5)",      avg_rating if avg_rating else "—")

    # ══════════════════════════════════════════════════════════════════════════
    # Лист 2 — Детали
    # ══════════════════════════════════════════════════════════════════════════
    ws2 = wb.create_sheet("Детали")

    cols = [
        ("№",                 5),
        ("Дата / время",     18),
        ("Пользователь",     18),
        ("Модуль",           15),
        ("Статус",           12),
        ("Тип документа",    22),
        ("Оценка",           10),
        ("Комментарий",      40),
        ("ID задания",       38),
    ]
    for col_idx, (title, width) in enumerate(cols, 1):
        ws2.column_dimensions[get_column_letter(col_idx)].width = width
        cell = ws2.cell(row=1, column=col_idx, value=title)
        cell.font = Font(bold=True, color="FFFFFF", size=10)
        cell.fill = PatternFill("solid", fgColor=CLR_HEADER)
        cell.alignment = Alignment(horizontal="center")
        cell.border = _thin_border()

    from app.models.models import Document as DocModel
    docs_map: dict = {}
    if job_ids:
        for doc in db.query(DocModel).filter(DocModel.job_id.in_(job_ids)).all():
            docs_map.setdefault(doc.job_id, []).append(doc.document_type)

    DOC_LABELS = {
        "UPD":             "Универсальный передаточный документ",
        "Act":             "Акт выполненных работ",
        "Invoice":         "Счет на оплату",
        "Invoice-Factura": "Счет-фактура",
        "unknown":         "Неизвестный тип",
    }

    for row_idx, job in enumerate(sorted(jobs, key=lambda j: j.created_at), 2):
        fb      = feedbacks.get(job.id)
        user    = users_map.get(job.user_id)
        dtypes  = ", ".join(DOC_LABELS.get(d, d) for d in docs_map.get(job.id, []))
        rating_str = ("👍 Хорошо" if fb and fb.rating >= 4 else
                      "👎 Плохо"  if fb and fb.rating <= 2 else
                      "—")

        row_fill = None
        if fb and fb.rating >= 4:
            row_fill = PatternFill("solid", fgColor=CLR_GOOD)
        elif fb and fb.rating <= 2:
            row_fill = PatternFill("solid", fgColor=CLR_BAD)

        values = [
            row_idx - 1,
            job.created_at.strftime("%d.%m.%Y %H:%M") if job.created_at else "",
            user.username if user else str(job.user_id or "—"),
            job.module or "—",
            job.status,
            dtypes or "—",
            rating_str,
            fb.comment if fb and fb.comment else "",
            str(job.id),
        ]
        for col_idx, val in enumerate(values, 1):
            cell = ws2.cell(row=row_idx, column=col_idx, value=val)
            cell.border = _thin_border()
            cell.font = Font(size=9)
            if row_fill:
                cell.fill = row_fill
            if col_idx == 8:  # комментарий — перенос
                cell.alignment = Alignment(wrap_text=True)

    # ── Итоговая строка на листе 2 ────────────────────────────────────────────
    last = len(jobs) + 2
    ws2.cell(row=last, column=1, value="ИТОГО").font = Font(bold=True, size=10)
    ws2.cell(row=last, column=2, value=f"{len(jobs)} сканирований").font = Font(bold=True, size=10)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def send_daily_report():
    """Точка входа для планировщика. Собирает отчёт и шлёт письмо."""
    if not settings.SMTP_HOST:
        logger.info("Ежедневный отчёт: SMTP_HOST не задан — пропускаем.")
        return

    db = SessionLocal()
    try:
        report_date = datetime.now(timezone.utc) - timedelta(days=1)
        xlsx_bytes  = build_daily_report(db, report_date)
        filename    = f"OCR_Report_{report_date.strftime('%Y-%m-%d')}.xlsx"

        # Получатели: из БД (приоритет) или из env
        db_recipients = get_setting(db, "report_recipients")
        raw = db_recipients if db_recipients else settings.REPORT_EMAIL_TO
        recipients = [r.strip() for r in raw.split(",") if r.strip()]

        if not recipients:
            logger.info("Ежедневный отчёт: список получателей пуст — пропускаем.")
            return

        msg = MIMEMultipart()
        msg["From"]    = settings.SMTP_FROM or settings.SMTP_USER
        msg["To"]      = ", ".join(recipients)
        msg["Subject"] = f"[OCR Engine] Ежедневный отчёт за {report_date.strftime('%d.%m.%Y')}"

        body = (
            f"Добрый день!\n\n"
            f"Во вложении ежедневный отчёт OCR Engine за {report_date.strftime('%d.%m.%Y')}.\n\n"
            f"• Лист «Сводка» — общая статистика: сканирования, токены, расходы, оценки.\n"
            f"• Лист «Детали» — все задания за день: пользователь, тип документа, оценка, комментарий.\n\n"
            f"С уважением,\nOCR Engine"
        )
        msg.attach(MIMEText(body, "plain", "utf-8"))

        part = MIMEApplication(xlsx_bytes, Name=filename)
        part["Content-Disposition"] = f'attachment; filename="{filename}"'
        msg.attach(part)

        if settings.SMTP_USE_TLS:
            server = smtplib.SMTP(settings.SMTP_HOST, settings.SMTP_PORT, timeout=30)
            server.starttls()
        else:
            server = smtplib.SMTP_SSL(settings.SMTP_HOST, settings.SMTP_PORT, timeout=30)

        if settings.SMTP_USER and settings.SMTP_PASSWORD:
            server.login(settings.SMTP_USER, settings.SMTP_PASSWORD)

        server.sendmail(msg["From"], recipients, msg.as_string())
        server.quit()
        logger.info(f"Ежедневный отчёт отправлен на: {', '.join(recipients)}")

    except Exception as e:
        logger.error(f"Ошибка отправки ежедневного отчёта: {e}")
    finally:
        db.close()
